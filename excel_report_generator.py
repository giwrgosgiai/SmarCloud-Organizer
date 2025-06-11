#!/usr/bin/env python3
"""
📊 ADVANCED EXCEL REPORT GENERATOR
🏢 Enterprise-grade Excel reporting with interactive features
🇬🇷 Full Greek support with dropdown menus and conditional formatting
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
from typing import Dict, List, Any
from pathlib import Path

class ExcelReportGenerator:
    """Advanced Excel report generator with interactive features"""

    def __init__(self):
        self.setup_styles()
        self.writer = None

    def setup_styles(self):
        """Setup predefined styles for Excel formatting"""
        self.styles = {
            'header': {
                'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
            },
            'data': {
                'font': Font(name='Calibri', size=10),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(border_style='thin', color='D3D3D3'),
                    right=Side(border_style='thin', color='D3D3D3'),
                    top=Side(border_style='thin', color='D3D3D3'),
                    bottom=Side(border_style='thin', color='D3D3D3')
                )
            },
            'summary': {
                'font': Font(name='Calibri', size=11, bold=True),
                'fill': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'issue': {
                'font': Font(name='Calibri', size=10, color='CC0000'),
                'fill': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
            },
            'success': {
                'font': Font(name='Calibri', size=10, color='006600'),
                'fill': PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center')
            }
        }

    def generate_comprehensive_report(self, audit_results: Dict[str, Any], output_file: str) -> str:
        """
        Generate a comprehensive Excel report from audit results.

        Args:
            audit_results: Dictionary containing audit results
            output_file: Path to save the Excel file

        Returns:
            Path to the generated Excel file
        """
        self.writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

        # Create Summary sheet
        self._create_summary_sheet(audit_results)

        # Create Files sheet
        self._create_files_sheet(audit_results)

        # Create Issues sheet
        self._create_issues_sheet(audit_results)

        # Save the Excel file
        self.writer.close()

        return output_file

    def _create_summary_sheet(self, audit_results: Dict[str, Any]):
        """Create the Summary sheet"""
        summary = audit_results['summary']

        # Handle timestamp gracefully
        ts = audit_results.get('timestamp', 'unknown')
        if ts and ts != 'unknown':
            try:
                ts_fmt = datetime.fromisoformat(ts).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                ts_fmt = 'unknown'
        else:
            ts_fmt = 'unknown'

        # Create summary data
        summary_data = {
            'Metric': [
                'Total Files Analyzed',
                'Total Issues Found',
                'Audit Date',
                'Target Directory'
            ],
            'Value': [
                summary['total_files'],
                summary['total_issues'],
                ts_fmt,
                audit_results['target_directory']
            ]
        }

        # Create document types data
        doc_types = summary['document_types']
        for doc_type, count in doc_types.items():
            summary_data['Metric'].append(f'Files of type: {doc_type}')
            summary_data['Value'].append(count)

        # Create issue types data
        issue_types = summary['issue_types']
        for issue_type, count in issue_types.items():
            summary_data['Metric'].append(f'Issues of type: {issue_type}')
            summary_data['Value'].append(count)

        # Create DataFrame and write to Excel
        df = pd.DataFrame(summary_data)
        df.to_excel(self.writer, sheet_name='Summary', index=False)

        # Get workbook and worksheet objects
        workbook = self.writer.book
        worksheet = self.writer.sheets['Summary']

        # Add formatting
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D9E1F2',
            'border': 1
        })

        # Apply header format
        worksheet.conditional_format(0, 0, 0, 1, {'type': 'no_blanks', 'format': header_format})

        # Adjust column widths
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 20)

    def _create_files_sheet(self, audit_results: Dict[str, Any]):
        """Create the Files sheet"""
        files_data = []

        for file_info in audit_results['files_analyzed']:
            # Handle modified_date gracefully
            mod_date = file_info.get('modified_date', 'unknown')
            if mod_date and mod_date != 'unknown':
                try:
                    mod_date_fmt = datetime.fromisoformat(mod_date).strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    mod_date_fmt = 'unknown'
            else:
                mod_date_fmt = 'unknown'
            files_data.append({
                'Filename': file_info['filename'],
                'Document Type': file_info['document_type'],
                'Size (bytes)': file_info.get('size', ''),
                'Modified Date': mod_date_fmt,
                'Full Path': file_info['full_path']
            })

        # Create DataFrame and write to Excel
        df = pd.DataFrame(files_data)
        df.to_excel(self.writer, sheet_name='Files', index=False)

        # Get workbook and worksheet objects
        workbook = self.writer.book
        worksheet = self.writer.sheets['Files']

        # Add formatting
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D9E1F2',
            'border': 1
        })

        # Apply header format
        worksheet.conditional_format(0, 0, 0, len(df.columns)-1, {'type': 'no_blanks', 'format': header_format})

        # Adjust column widths
        worksheet.set_column('A:A', 30)  # Filename
        worksheet.set_column('B:B', 20)  # Document Type
        worksheet.set_column('C:C', 15)  # Size
        worksheet.set_column('D:D', 20)  # Modified Date
        worksheet.set_column('E:E', 50)  # Full Path

    def _create_issues_sheet(self, audit_results: Dict[str, Any]):
        """Create the Issues sheet"""
        issues_data = []

        for issue in audit_results['issues_found']:
            issues_data.append({
                'Issue Type': issue['type'],
                'Severity': issue['severity'],
                'Message': issue['message'],
                'Filename': issue['file_info']['filename'],
                'Document Type': issue['file_info']['document_type']
            })

        # Create DataFrame and write to Excel
        df = pd.DataFrame(issues_data)
        df.to_excel(self.writer, sheet_name='Issues', index=False)

        # Get workbook and worksheet objects
        workbook = self.writer.book
        worksheet = self.writer.sheets['Issues']

        # Add formatting
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D9E1F2',
            'border': 1
        })

        # Apply header format
        worksheet.conditional_format(0, 0, 0, len(df.columns)-1, {'type': 'no_blanks', 'format': header_format})

        # Add conditional formatting for severity
        warning_format = workbook.add_format({'bg_color': '#FFEB9C'})
        error_format = workbook.add_format({'bg_color': '#FFC7CE'})
        info_format = workbook.add_format({'bg_color': '#C6EFCE'})

        # Apply conditional formatting
        worksheet.conditional_format(1, 1, len(df), 1, {
            'type': 'text',
            'criteria': 'containing',
            'value': 'warning',
            'format': warning_format
        })
        worksheet.conditional_format(1, 1, len(df), 1, {
            'type': 'text',
            'criteria': 'containing',
            'value': 'error',
            'format': error_format
        })
        worksheet.conditional_format(1, 1, len(df), 1, {
            'type': 'text',
            'criteria': 'containing',
            'value': 'info',
            'format': info_format
        })

        # Adjust column widths
        worksheet.set_column('A:A', 20)  # Issue Type
        worksheet.set_column('B:B', 10)  # Severity
        worksheet.set_column('C:C', 50)  # Message
        worksheet.set_column('D:D', 30)  # Filename
        worksheet.set_column('E:E', 20)  # Document Type

    def create_folders_sheet(self, sheet, audit_data: Dict):
        """Create unknown folders analysis sheet"""

        unknown_folders = audit_data.get('unknown_folders', [])

        headers = ['Folder Path', 'Folder Name', 'Is Problematic', 'File Count', 'Recommendation', 'Selected Action', 'Notes']

        # Add headers
        for i, header in enumerate(headers):
            cell = sheet[f'{chr(65+i)}1']
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['header']['border']

        # Add folder data
        row = 2
        for folder in unknown_folders:
            row_data = [
                folder.get('folder_path', ''),
                folder.get('folder_name', ''),
                '✅' if folder.get('is_problematic', False) else '❌',
                folder.get('file_count', 0),
                folder.get('recommendation', ''),
                '',  # Empty for user selection
                'Requires manual review'
            ]

            for i, value in enumerate(row_data):
                cell = sheet[f'{chr(65+i)}{row}']
                cell.value = value

                if folder.get('is_problematic', False) and i < 5:
                    cell.font = self.styles['issue']['font']
                    cell.fill = self.styles['issue']['fill']
                else:
                    cell.font = self.styles['data']['font']

                cell.alignment = self.styles['data']['alignment']
                cell.border = self.styles['data']['border']

            row += 1

        # Add action dropdown
        action_dropdown = DataValidation(
            type="list",
            formula1='"Ignore,Merge with supplier folder,Rename to standard format,Move contents to proper location"',
            allow_blank=True
        )
        sheet.add_data_validation(action_dropdown)
        action_dropdown.add(f'F2:F{row-1}')

        self.auto_adjust_columns(sheet)

    def create_duplicates_sheet(self, sheet, audit_data: Dict):
        """Create duplicates analysis sheet"""

        duplicates = audit_data.get('duplicates', [])

        headers = ['File Path', 'Duplicate Of', 'Duplicate Type', 'File Size', 'Modified Date', 'Recommendation', 'Action', 'Notes']

        # Add headers
        for i, header in enumerate(headers):
            cell = sheet[f'{chr(65+i)}1']
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['header']['border']

        # Add duplicate data
        row = 2
        for duplicate in duplicates:
            file_path = Path(duplicate.get('file_path', ''))

            # Get file info
            try:
                file_size = file_path.stat().st_size / 1024  # KB
                mod_date = datetime.datetime.fromtimestamp(file_path.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
            except:
                file_size = 0
                mod_date = 'Unknown'

            row_data = [
                str(file_path),
                duplicate.get('duplicate_of', ''),
                duplicate.get('duplicate_type', 'exact_content'),
                f'{file_size:.1f} KB',
                mod_date,
                duplicate.get('recommendation', 'DELETE'),
                '',  # For user action
                'Detected duplicate content'
            ]

            for i, value in enumerate(row_data):
                cell = sheet[f'{chr(65+i)}{row}']
                cell.value = value
                cell.font = self.styles['data']['font']

                if duplicate.get('recommendation') == 'DELETE' and i == 5:
                    cell.fill = self.styles['issue']['fill']

                cell.alignment = self.styles['data']['alignment']
                cell.border = self.styles['data']['border']

            row += 1

        # Add action dropdown
        action_dropdown = DataValidation(
            type="list",
            formula1='"DELETE,KEEP,RENAME,MOVE_TO_BACKUP"',
            allow_blank=True
        )
        sheet.add_data_validation(action_dropdown)
        action_dropdown.add(f'G2:G{row-1}')

        self.auto_adjust_columns(sheet)

    def create_completeness_sheet(self, sheet, audit_data: Dict):
        """Create document completeness analysis sheet"""

        completeness = audit_data.get('completeness_analysis', {})

        headers = ['Container', 'Invoice Count', 'CE Count', 'Packing List', 'BL/HBL', 'Manual', 'Bank Proof', 'Completeness Score', 'Issues', 'Status']

        # Add headers
        for i, header in enumerate(headers):
            cell = sheet[f'{chr(65+i)}1']
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['header']['border']

        # Add completeness data
        row = 2
        for container, data in completeness.items():
            counts = data.get('document_counts', {})
            score = data.get('completeness_score', 0)
            issues = '; '.join(data.get('issues', []))

            status = 'COMPLETE' if score >= 80 else 'INCOMPLETE' if score >= 50 else 'CRITICAL'

            row_data = [
                container,
                counts.get('invoice', 0),
                counts.get('ce', 0),
                counts.get('packing_list', 0),
                counts.get('bl', 0),
                counts.get('manual', 0),
                counts.get('bank_proof', 0),
                f'{score:.1f}%',
                issues if issues else 'No issues',
                status
            ]

            for i, value in enumerate(row_data):
                cell = sheet[f'{chr(65+i)}{row}']
                cell.value = value

                # Color coding based on completeness score
                if status == 'CRITICAL':
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                elif status == 'INCOMPLETE':
                    cell.fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')

                cell.font = self.styles['data']['font']
                cell.alignment = self.styles['data']['alignment']
                cell.border = self.styles['data']['border']

            row += 1

        self.auto_adjust_columns(sheet)

    def create_recommendations_sheet(self, sheet, audit_data: Dict):
        """Create actionable recommendations sheet"""

        # Title
        sheet['A1'] = '💡 ACTIONABLE RECOMMENDATIONS'
        sheet['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
        sheet.merge_cells('A1:D1')

        row = 3

        # High priority recommendations
        recommendations = [
            {
                'category': 'HIGH PRIORITY',
                'items': [
                    'Review container mismatches - files may be in wrong folders',
                    'Delete duplicate files to save space and reduce confusion',
                    'Organize unknown/problematic folders',
                    'Source missing CE certificates for incomplete containers'
                ]
            },
            {
                'category': 'MEDIUM PRIORITY',
                'items': [
                    'Implement consistent naming convention for files',
                    'Create backup of important documents before reorganization',
                    'Extract shipping dates from BL/HBL documents',
                    'Calculate customs costs from available documents'
                ]
            },
            {
                'category': 'LOW PRIORITY',
                'items': [
                    'Archive old temporary files',
                    'Create document templates for future use',
                    'Implement automated filing system',
                    'Regular audit schedule (monthly/quarterly)'
                ]
            }
        ]

        for recommendation in recommendations:
            # Category header
            sheet[f'A{row}'] = recommendation['category']
            sheet[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color='CC0000' if 'HIGH' in recommendation['category'] else '996600' if 'MEDIUM' in recommendation['category'] else '006600')
            row += 1

            # Items
            for item in recommendation['items']:
                sheet[f'B{row}'] = f"• {item}"
                sheet[f'B{row}'].font = Font(name='Calibri', size=11)
                sheet[f'B{row}'].alignment = Alignment(wrap_text=True)
                row += 1

            row += 1  # Empty row between categories

        # Implementation checklist
        row += 2
        sheet[f'A{row}'] = 'IMPLEMENTATION CHECKLIST'
        sheet[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
        row += 1

        checklist_items = [
            'Backup all files before making changes',
            'Start with high-priority items',
            'Test changes on a small subset first',
            'Document all changes made',
            'Schedule follow-up audit'
        ]

        for item in checklist_items:
            sheet[f'A{row}'] = '☐'
            sheet[f'B{row}'] = item
            sheet[f'B{row}'].font = Font(name='Calibri', size=11)
            row += 1

        self.auto_adjust_columns(sheet)

    def auto_adjust_columns(self, sheet):
        """Auto-adjust column widths"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width